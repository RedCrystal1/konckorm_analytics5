import io
import logging
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("apps.reports")


HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def generate_overdue_registry_excel(date_from=None, date_to=None, counterparty=None):
    """Генерация Excel-отчёта: реестр просроченной задолженности."""
    from apps.registers.models import DebtByTerms

    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр просрочки"

    # Заголовок
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Реестр просроченной задолженности на {date.today().strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(bold=True, size=14)

    # Шапка таблицы
    headers = [
        "Контрагент",
        "Документ",
        "Дата документа",
        "Срок оплаты",
        "Сумма (руб.)",
        "Просрочка (руб.)",
        "Дней просрочки",
        "Статус",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    _style_header(ws, 3, len(headers))

    # Данные
    qs = DebtByTerms.objects.exclude(
        status=DebtByTerms.DebtStatus.CURRENT
    ).select_related("counterparty", "source_document").order_by("-overdue_days")

    if date_from:
        qs = qs.filter(planned_payment_date__gte=date_from)
    if date_to:
        qs = qs.filter(planned_payment_date__lte=date_to)
    if counterparty:
        qs = qs.filter(counterparty=counterparty)
    for row_idx, record in enumerate(qs, start=4):
        ws.cell(row=row_idx, column=1, value=record.counterparty.name)
        ws.cell(row=row_idx, column=2, value=str(record.source_document))
        ws.cell(
            row=row_idx,
            column=3,
            value=record.source_document.date.strftime("%d.%m.%Y"),
        )
        ws.cell(
            row=row_idx,
            column=4,
            value=record.planned_payment_date.strftime("%d.%m.%Y"),
        )
        ws.cell(row=row_idx, column=5, value=float(record.amount_rub))
        ws.cell(row=row_idx, column=6, value=float(record.overdue_amount))
        ws.cell(row=row_idx, column=7, value=record.overdue_days)
        ws.cell(row=row_idx, column=8, value=record.get_status_display())

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    # Ширина столбцов
    col_widths = [35, 30, 14, 14, 16, 16, 16, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def generate_payment_calendar_excel(date_from=None, date_to=None, counterparty=None):
    """Генерация Excel-отчёта: платёжный календарь."""
    from apps.registers.models import PlannedPayment

    wb = Workbook()
    ws = wb.active
    ws.title = "Платёжный календарь"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Платёжный календарь"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Дата платежа", "Контрагент", "Договор", "Сумма", "Приоритет", "Статус"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    _style_header(ws, 3, len(headers))

    qs = PlannedPayment.objects.select_related(
        "counterparty", "contract"
    ).order_by("planned_date")

    if date_from:
        qs = qs.filter(planned_date__gte=date_from)
    if date_to:
        qs = qs.filter(planned_date__lte=date_to)
    if counterparty:
        qs = qs.filter(counterparty=counterparty)
    for row_idx, pp in enumerate(qs, start=4):
        ws.cell(row=row_idx, column=1, value=pp.planned_date.strftime("%d.%m.%Y"))
        ws.cell(row=row_idx, column=2, value=pp.counterparty.name)
        ws.cell(row=row_idx, column=3, value=str(pp.contract or "—"))
        ws.cell(row=row_idx, column=4, value=float(pp.amount))
        ws.cell(row=row_idx, column=5, value=pp.get_priority_display())
        ws.cell(row=row_idx, column=6, value=pp.get_status_display())

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    col_widths = [14, 35, 25, 16, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def generate_procurement_structure_excel(date_from=None, date_to=None, counterparty=None):
    """Генерация Excel-отчёта: структура закупок."""
    from apps.registers.models import ProcurementVolume

    wb = Workbook()
    ws = wb.active
    ws.title = "Структура закупок"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Структура закупок по поставщикам"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Поставщик", "Период", "Вид закупок", "Объём (руб.)", "Доля (%)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    _style_header(ws, 3, len(headers))

    qs = ProcurementVolume.objects.select_related("counterparty").order_by("-volume_rub")

    if date_from:
        qs = qs.filter(period_start__gte=date_from)
    if date_to:
        qs = qs.filter(period_end__lte=date_to)
    if counterparty:
        qs = qs.filter(counterparty=counterparty)
    for row_idx, pv in enumerate(qs, start=4):
        ws.cell(row=row_idx, column=1, value=pv.counterparty.name)
        ws.cell(
            row=row_idx,
            column=2,
            value=f"{pv.period_start.strftime('%d.%m.%Y')} — {pv.period_end.strftime('%d.%m.%Y')}",
        )
        ws.cell(row=row_idx, column=3, value=pv.get_procurement_kind_display())
        ws.cell(row=row_idx, column=4, value=float(pv.volume_rub))
        ws.cell(row=row_idx, column=5, value=float(pv.share_percent))

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    col_widths = [35, 25, 16, 18, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def generate_debt_by_terms_excel(date_from=None, date_to=None, counterparty=None):
    """Генерация Excel: задолженность по срокам."""
    from apps.registers.models import DebtByTerms

    wb = Workbook()
    ws = wb.active
    ws.title = "Задолженность по срокам"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Задолженность по срокам на {date.today().strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(bold=True, size=14)

    headers = [
        "Контрагент", "Договор", "Документ", "Дата документа",
        "Срок оплаты", "Сумма (руб.)", "Дней просрочки", "Статус",
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=header)
    _style_header(ws, 3, len(headers))

    qs = DebtByTerms.objects.select_related(
        "counterparty", "contract", "source_document"
    ).order_by("-overdue_days")

    if date_from:
        qs = qs.filter(planned_payment_date__gte=date_from)
    if date_to:
        qs = qs.filter(planned_payment_date__lte=date_to)
    if counterparty:
        qs = qs.filter(counterparty=counterparty)

    for row_idx, r in enumerate(qs, start=4):
        ws.cell(row=row_idx, column=1, value=r.counterparty.name)
        ws.cell(row=row_idx, column=2, value=str(r.contract or "—"))
        ws.cell(row=row_idx, column=3, value=str(r.source_document))
        ws.cell(row=row_idx, column=4, value=r.source_document.date.strftime("%d.%m.%Y"))
        ws.cell(row=row_idx, column=5, value=r.planned_payment_date.strftime("%d.%m.%Y"))
        ws.cell(row=row_idx, column=6, value=float(r.amount_rub))
        ws.cell(row=row_idx, column=7, value=r.overdue_days)
        ws.cell(row=row_idx, column=8, value=r.get_status_display())
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    col_widths = [35, 25, 25, 14, 14, 16, 16, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_counterparty_card_excel(counterparty=None, date_from=None, date_to=None):
    """Генерация Excel: карточка контрагента."""
    from apps.counterparties.models import Counterparty
    from apps.documents.models import GoodsReceipt, PaymentOrder

    if not counterparty:
        raise ValueError("Для отчёта «Карточка контрагента» необходимо выбрать контрагента.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Карточка контрагента"

    # Заголовок
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Карточка контрагента: {counterparty.name}"
    ws["A1"].font = Font(bold=True, size=14)

    # Реквизиты
    info = [
        ("Наименование", counterparty.name),
        ("Полное наименование", counterparty.full_name or "—"),
        ("ИНН", counterparty.inn),
        ("КПП", counterparty.kpp or "—"),
        ("Телефон", counterparty.phone or "—"),
        ("Email", counterparty.email or "—"),
        ("Контактное лицо", counterparty.contact_person or "—"),
        ("Ключевой поставщик", "Да" if counterparty.is_key_supplier else "Нет"),
    ]
    for row_idx, (label, value) in enumerate(info, start=3):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    # Документы
    start_row = len(info) + 5
    ws.cell(row=start_row, column=1, value="Поступления").font = Font(bold=True, size=12)
    headers = ["Номер", "Дата", "Сумма", "Оплачено", "Остаток", "Срок оплаты", "Оплачен"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row + 1, column=col, value=h)
    _style_header(ws, start_row + 1, len(headers))

    receipts = GoodsReceipt.objects.filter(counterparty=counterparty).order_by("-date")
    if date_from:
        receipts = receipts.filter(date__gte=date_from)
    if date_to:
        receipts = receipts.filter(date__lte=date_to)

    for row_idx, r in enumerate(receipts, start=start_row + 2):
        ws.cell(row=row_idx, column=1, value=r.number)
        ws.cell(row=row_idx, column=2, value=r.date.strftime("%d.%m.%Y"))
        ws.cell(row=row_idx, column=3, value=float(r.amount))
        ws.cell(row=row_idx, column=4, value=float(r.paid_amount))
        ws.cell(row=row_idx, column=5, value=float(r.outstanding_amount))
        ws.cell(row=row_idx, column=6, value=r.payment_due_date.strftime("%d.%m.%Y"))
        ws.cell(row=row_idx, column=7, value="Да" if r.is_paid else "Нет")
        for col in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col).border = THIN_BORDER

    col_widths = [15, 12, 14, 14, 14, 14, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
