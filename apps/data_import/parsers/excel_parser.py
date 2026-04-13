import logging

import openpyxl

logger = logging.getLogger("apps.data_import")


def parse_excel_counterparties(file_obj):
    """Парсинг листа 'Контрагенты' из Excel."""
    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True)

    if "Контрагенты" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["Контрагенты"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    results = []
    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue

        try:
            data = {
                "code_1c": str(row[0]).strip() if row[0] else "",
                "name": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                "full_name": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                "inn": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                "kpp": str(row[4]).strip() if len(row) > 4 and row[4] else "",
                "phone": str(row[5]).strip() if len(row) > 5 and row[5] else "",
                "email": str(row[6]).strip() if len(row) > 6 and row[6] else "",
                "contact_person": str(row[7]).strip() if len(row) > 7 and row[7] else "",
            }
            results.append({"row": i, "data": data})
        except Exception as e:
            logger.warning("Ошибка парсинга строки %d: %s", i, e)
            results.append({"row": i, "error": str(e)})

    return results


def parse_excel_contracts(file_obj):
    """Парсинг листа 'Договоры' из Excel."""
    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, read_only=True)

    if "Договоры" not in wb.sheetnames:
        wb.close()
        return []

    ws = wb["Договоры"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    results = []
    for i, row in enumerate(rows, start=2):
        if not row or not row[0]:
            continue

        try:
            data = {
                "code_1c": str(row[0]).strip() if row[0] else "",
                "counterparty_code": str(row[1]).strip() if len(row) > 1 and row[1] else "",
                "number": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                "date": row[3] if len(row) > 3 else None,
                "kind": str(row[4]).strip() if len(row) > 4 and row[4] else "supply",
                "payment_days": int(row[5]) if len(row) > 5 and row[5] else 30,
            }
            results.append({"row": i, "data": data})
        except Exception as e:
            logger.warning("Ошибка парсинга строки %d: %s", i, e)
            results.append({"row": i, "error": str(e)})

    return results
